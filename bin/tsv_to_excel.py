#!/usr/bin/env python3

import glob
import os
import re
import sys
from argparse import ArgumentParser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Color


def parseArgs():
    parser = ArgumentParser(
        description="Create an Excel files from TSV files and optionally color sheet tabs.",
        add_help=False,
        epilog='NOTE: sheet tabs are named based on filenames with .tsv removed and Summary*. removed')
    req = parser.add_argument_group('Required')
    req.add_argument('tsv_files',
        nargs='+',
        metavar='FILE',
        type=str,
        help='One or more TSV files (or wildcard pattern) to convert to single XLSX.')
    opt = parser.add_argument_group('Optional')
    opt.add_argument('-h', '--help',
        action='help',
		help='show this help message and exit')
    opt.add_argument('-d', '--color-dict', 
        metavar='FILE',
        type=str,
        required=False,
        help='Path to the color dictionary file. Each line must have Summary*. prefix and .tsv suffix stripped, and colors are RGB, e.g., "MLST=255, 255, 0" [None]')
    opt.add_argument('-o', '--outfile',
        required=False,
        type=str,
        metavar='FILE',
        default='Summary-Report.xlsx', 
        help='Microsoft Excel XLSX formatted output file [Summary-Report.xlsx]')
    opt.add_argument('-c', '--xlsx-property-category',
        required=False,
        type=str,
        metavar='STR',
        default='',
        help='Output XLSX file property category [None]')
    opt.add_argument('-s', '--xlsx-property-subject',
        required=False,
        type=str,
        metavar='STR',
        default='',
        help='Output XLSX file property subject [None]')
    opt.add_argument('-t', '--xlsx-property-title',
        required=False,
        type=str,
        metavar='STR',
        default='',
        help='Output XLSX file property title [None]')
    return parser.parse_args()

def get_target_creation_time(file):
    """Get the creation time of the target file, following symlinks."""
    target = os.path.realpath(file)
    return os.path.getctime(target)

def adjust_tsv_headers(tsv_file):
    """Fix TSV files that have missing header column names and fill in missing data cells with empty."""
    with open(tsv_file, 'r') as f:
        lines = f.readlines()

    header = lines[0].strip().split('\t')
    data_rows = [line.strip().split('\t') for line in lines[1:]]

    max_cols = max(len(row) for row in data_rows)

    if len(header) < max_cols:
        header.extend([header[-1]] * (max_cols - len(header)))

    for i, row in enumerate(data_rows):
        if len(row) < max_cols:
            data_rows[i].extend([''] * (max_cols - len(row)))

    data = pd.DataFrame(data_rows, columns=header)
    return data

def create_summary_workbook(output_file, tsv_file):
    """Create an Excel workbook summary from a TSV file."""
    prefix_stripped = re.sub(r'^Summary.*?\.', '', tsv_file)
    sheet_name = prefix_stripped.removesuffix(".tsv")
    try:
        data = adjust_tsv_headers(tsv_file)
        data.to_excel(output_file, sheet_name=sheet_name, index=False)
    except pd.errors.ParserError as e:
        print(f"ERROR: Skipping {tsv_file} due to parsing error: {e}")

def rgb_to_hex(rgb):
    """Convert an RGB tuple to a hex string."""
    return ''.join(f'{x:02X}' for x in rgb)

def load_color_dict(color_dict_file):
    """Load the color dictionary from a text file."""
    color_dict = {}
    with open(color_dict_file, 'r') as ifh:
        for line in ifh:
            line = line.strip()

            # Skip commented lines
            if line.startswith('#') or line.startswith('//') or line.startswith(';'):
                continue

            # The RGB color has to be after the equals sign
            if '=' in line:
                name, rgb_str = line.split('=', 1)
                name = name.strip()
                rgb_str = rgb_str.strip().strip('{}')
                
                print(f"Processing XLSX sheet line: name={name}, rgb_str={rgb_str}")

                try:
                    rgb = tuple(map(int, rgb_str.split(',')))
                    if len(rgb) == 3:
                        color_dict[name] = rgb_to_hex(rgb)
                    else:
                        print(f"Warning: RGB value for '{name}' lacks 3 components.")
                except ValueError:
                    print(f"Error: Invalid RGB value '{rgb_str}' for '{name}'.")
    
    return color_dict

def color_sheet_tabs(filename, color_dict_hex, title, category, subject):
    """Color tabs in XLSX based on a color dictionary of names and color codes."""
    workbook = load_workbook(filename)

    workbook.properties.title = title
    workbook.properties.category = category
    workbook.properties.subject = subject

    for sheet_name in workbook.sheetnames:
        if sheet_name in color_dict_hex:
            hex_color = color_dict_hex[sheet_name]
            color = Color(rgb=hex_color)
            workbook[sheet_name].sheet_properties.tabColor = color

    workbook.save(filename)

def main():
    # I/O Handling
    opts = parseArgs()
    xlsx_outfile = opts.outfile

    _, ext = os.path.splitext(xlsx_outfile)
    if ext.lower() != ".xlsx":
        sys.stderr.write("ERROR: The file extension must be '.xlsx'.\n")
        sys.exit(1)

    # Ensure the directory exists
    output_dir = os.path.dirname(os.path.realpath(xlsx_outfile))
    os.makedirs(output_dir, exist_ok=True)

    tsv_files = []
    for pattern in opts.tsv_files:
        tsv_files.extend(glob.glob(pattern))

    if not tsv_files:
        print("No TSV files found.")
        return

    # Sort TSV files by creation time (Follow symlink targets)
    tsv_files.sort(key=get_target_creation_time)

    # Do the work, converting all TSvs into single XLSX
    with pd.ExcelWriter(xlsx_outfile, engine='openpyxl') as output_file:
        for tsv_file in tsv_files:
            create_summary_workbook(output_file, tsv_file)

    # Optionally color the XLSX tabs
    if opts.color_dict is not None:
        color_dict_hex = load_color_dict(opts.color_dict)
        color_sheet_tabs(xlsx_outfile, color_dict_hex, opts.xlsx_property_title, opts.xlsx_property_category, opts.xlsx_property_subject)

if __name__ == "__main__":
    main()
